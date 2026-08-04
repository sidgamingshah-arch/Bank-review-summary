"""Code-list masters (segment/relationship/etc. as governed code-value masters),
the dropdown-source endpoints, template referential validation, and per-master
Excel download/upload."""
from __future__ import annotations

import io

import openpyxl
from fastapi.testclient import TestClient

import cam.services.master_config.main as mc
from cam.services.master_config import xlsx_io
from cam.services.master_config.schemas import validate_payload
from tests.test_master_config import publish

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _create_publish(c, admin, admin2, seg, key, payload):
    assert c.post(f"/api/masters/{seg}", json={"key": key, "payload": payload},
                  headers=admin).status_code == 201, key
    publish(c, seg, key, admin, admin2)


def test_codelist_master_and_dropdown_endpoint(admin_headers, admin2_headers):
    # 'risk_grade' is inert w.r.t. template validation (only segment/relationship
    # feed that), so this stays isolated from the shared test DB.
    with TestClient(mc.app) as c:
        _create_publish(c, admin_headers, admin2_headers, "codelists", "risk_grade", {
            "name": "risk_grade", "description": "Internal risk grade",
            "entries": [{"code": "aaa", "label": "AAA", "active": True, "order": 2},
                        {"code": "bbb", "label": "BBB", "active": True, "order": 1},
                        {"code": "old", "label": "Retired", "active": False, "order": 3}]})
        r = c.get("/api/masters/published/codelist/risk_grade", headers=admin_headers)
        assert r.status_code == 200
        entries = r.json()["entries"]
        # active only, sorted by order
        assert [e["code"] for e in entries] == ["bbb", "aaa"]
        assert entries[0]["label"] == "BBB"


def test_template_segment_relationship_validated_against_codelists():
    # unit-level (no shared DB): the referential check fires only once the code
    # list is populated, and rejects codes outside it.
    base = {"name": "Corp CAM", "template_instructions": "",
            "sections": [{"order": 1, "section_code": "s1"}], "required_doc_types": []}
    cat = {"doctype_codes": set(), "prompt_keys": {"s1"}, "industry_codes": set()}

    # no code lists yet -> lenient (any value accepted)
    _, errs = validate_payload("template", "t1", {**base, "segment": "x", "relationship": "y"}, **cat)
    assert errs == []

    # populated lists -> enforced
    strict = {**cat, "segment_codes": {"corporate"}, "relationship_codes": {"etb"}}
    _, errs = validate_payload("template", "t1", {**base, "segment": "corporate", "relationship": "etb"}, **strict)
    assert errs == []
    _, errs = validate_payload("template", "t1", {**base, "segment": "bogus", "relationship": "etb"}, **strict)
    assert any("segment" in e for e in errs)


def test_published_sections_endpoint(admin_headers, admin2_headers):
    with TestClient(mc.app) as c:
        _create_publish(c, admin_headers, admin2_headers, "prompts", "cl_section", {
            "section_code": "cl_section", "section_name": "CL Section", "scope": "section",
            "prompt_text": "Assess {{borrower_name}}.", "source_doc_types": [],
            "uses_industry_kpis": False})
        out = c.get("/api/masters/published/sections", headers=admin_headers).json()
        codes = [s["code"] for s in out]
        assert "cl_section" in codes
        assert "global_standing_rules" not in codes  # scope filter
        assert next(s for s in out if s["code"] == "cl_section")["name"] == "CL Section"


def test_per_type_excel_template_download(admin_headers):
    with TestClient(mc.app) as c:
        for seg in ("codelists", "doctypes", "templates"):
            r = c.get(f"/api/masters/{seg}/xlsx-template", headers=admin_headers)
            assert r.status_code == 200 and r.headers["content-type"] == XLSX
            wb = openpyxl.load_workbook(io.BytesIO(r.content))
            if seg == "templates":  # spans two linked sheets
                assert {"templates", "template_sections"} <= set(wb.sheetnames)
            else:
                assert seg in wb.sheetnames


def test_per_type_excel_upload_creates_drafts(admin_headers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "codelists"
    ws.append(xlsx_io.CODELIST_COLS)  # list_key,list_name,description,code,label,active,order
    ws.append(["tenor_band", "Tenor band", "Loan tenor", "short", "Short (<1y)", True, 1])
    ws.append(["tenor_band", "Tenor band", "Loan tenor", "long", "Long (>3y)", True, 2])
    buf = io.BytesIO()
    wb.save(buf)
    with TestClient(mc.app) as c:
        r = c.post("/api/masters/codelists/xlsx-upload",
                   files={"file": ("cl.xlsx", buf.getvalue(), XLSX)}, headers=admin_headers)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["errors"] == []
        assert any("tenor_band" in str(x) for x in body["created"])
        # landed as a draft (maker-checker unchanged)
        v = c.get("/api/masters/codelists/tenor_band/versions/1", headers=admin_headers).json()
        assert v["status"] == "draft"
        assert {e["code"] for e in v["payload"]["entries"]} == {"short", "long"}
