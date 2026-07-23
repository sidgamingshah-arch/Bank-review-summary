"""Bulk masters upload via the Excel template: download, round-trip, drafts
under maker-checker, dependency ordering, and per-row error reporting."""
from __future__ import annotations

import io

import openpyxl
from fastapi.testclient import TestClient

from cam.services.master_config import xlsx_io
from cam.services.master_config.main import XLSX_MEDIA, app, engine
from cam.services.master_config.models import Base

Base.metadata.create_all(engine)  # order-independent: ensure tables exist
client = TestClient(app)


def _filled(rows_extra: dict | None = None) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_io.build_template_workbook()))
    wb["doctypes"].append(["bulk_af", "Bulk AF", "desc", "annual report",
                           "balance sheet", True, "pdf", 25, 10, ""])
    wb["industries"].append(["mfg", "Manufacturing", "bulk_steel", "Steel"])
    wb["prompts"].append(["bulk_sec", "Bulk Section", "section",
                          "Draft for {{borrower_name}} using {{doc:bulk_af}}.",
                          "bulk_af", False, True, "", "", "", ""])
    wb["kpi_sets"].append(["bulk_steel", "k1", "KPI One", "def", "x",
                           "higher_better", "10", "bulk_sec"])
    wb["templates"].append(["bulk_tpl", "Bulk Tpl", "corporate", "etb",
                            "UK English", "bulk_af"])
    wb["template_sections"].append(["bulk_tpl", 1, "bulk_sec", True, "", "200 words", False])
    for sheet, row in (rows_extra or {}).items():
        wb[sheet].append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(data: bytes, headers) -> dict:
    r = client.post("/api/masters/bulk-upload", headers=headers,
                    files={"file": ("masters.xlsx", data, XLSX_MEDIA)})
    assert r.status_code == 200, r.text
    return r.json()


def test_template_download(admin_headers):
    r = client.get("/api/masters/bulk-template", headers=admin_headers)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    for sheet in ("README", "doctypes", "industries", "prompts", "kpi_sets",
                  "templates", "template_sections"):
        assert sheet in wb.sheetnames


def test_bulk_upload_creates_drafts_in_dependency_order(admin_headers):
    body = _upload(_filled(), admin_headers)
    assert body["errors"] == [], body["errors"]
    created = {c["entry"] for c in body["created"]}
    assert {"doctype:bulk_af", "industry:bulk_steel", "prompt:bulk_sec",
            "kpi_set:bulk_steel", "template:bulk_tpl"} <= created
    # everything lands as a draft — never auto-published
    for seg, key in [("doctypes", "bulk_af"), ("templates", "bulk_tpl"),
                     ("prompts", "bulk_sec")]:
        item = client.get(f"/api/masters/{seg}/{key}", headers=admin_headers).json()
        assert item["published_version"] is None
    # the opt-in flag survived the round-trip through the schema
    v = client.get("/api/masters/prompts/bulk_sec/versions/1", headers=admin_headers).json()
    assert v["payload"]["uses_external_context"] is True


def test_bad_row_reported_others_still_import(admin_headers):
    # a second template referencing a prompt that does not exist -> one error,
    # the valid entries still import
    extra = {"templates": ["bulk_tpl2", "Bad Tpl", "corporate", "etb", "", ""],
             "template_sections": ["bulk_tpl2", 1, "nonexistent_prompt", True, "", "", False]}
    body = _upload(_filled(extra), admin_headers)
    assert any("bulk_tpl2" in e.get("entry", "") for e in body["errors"])
    entries = {c["entry"] for c in body["created"]} | {c["entry"] for c in body["updated"]}
    assert "template:bulk_tpl" in entries


def test_unreadable_file_reports_error(admin_headers):
    body = _upload(b"not a workbook", admin_headers)
    assert body["errors"] and "readable" in body["errors"][0]["message"]
    assert body["created"] == [] and body["updated"] == []


def test_numeric_overflow_is_per_row_error_not_500(admin_headers):
    # an over-large int cell must degrade to a row error, not crash the upload
    extra = {"doctypes": ["bulk_ovf", "Overflow", "d", "", "", True, "", "1e400", 10, ""]}
    body = _upload(_filled(extra), admin_headers)
    assert any(e.get("sheet") == "doctypes" for e in body["errors"])
    # valid rows still imported (created on first run, updated on re-runs — shared DB)
    entries = {c["entry"] for c in body["created"]} | {c["entry"] for c in body["updated"]}
    assert "doctype:bulk_af" in entries


def test_missing_key_column_reports_sheet_error(admin_headers):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_io.build_template_workbook()))
    wb["doctypes"].cell(row=1, column=1).value = "Code"  # recased key header
    wb["doctypes"].append(["kc_dt", "Name", "d", "", "", True, "", 25, 10, ""])
    buf = io.BytesIO()
    wb.save(buf)
    body = _upload(buf.getvalue(), admin_headers)
    assert any(e.get("sheet") == "doctypes" and "key column" in e["message"]
               for e in body["errors"])


def test_orphaned_template_sections_reported(admin_headers):
    extra = {"template_sections": ["orphan_tpl", 1, "exec_summary", True, "", "", False]}
    body = _upload(_filled(extra), admin_headers)
    assert any(e.get("sheet") == "template_sections" and "orphan_tpl" in e["message"]
               for e in body["errors"])
