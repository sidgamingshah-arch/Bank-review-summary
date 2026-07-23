"""Bulk masters maintenance via a single Excel workbook (FR-A12, extended).

An admin downloads a template workbook — one sheet per master type, with a
worked example row and a README — fills it in, and uploads it. Every row lands
as a DRAFT under maker-checker (never published directly), reusing the same
import path as the JSON bundle.

Cell conventions
----------------
* list cells (synonyms, keywords, source_doc_types, sections, ...) are
  separated by ``|`` (pipe); commas and newlines are also accepted.
* boolean cells accept TRUE/FALSE, 1/0, yes/no.
* a row is IGNORED when its key cell is blank, starts with ``#``, or starts
  with ``example`` — so the shipped example rows are safe to leave in place.

Columns are matched by header NAME, so column order does not matter and unknown
columns are ignored. Structural/referential validation is deferred to
``schemas.validate_payload`` at import time; this module only shapes rows into
master payloads and reports rows it cannot even parse.
"""
from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Font

# ------------------------------------------------------------------ columns
DOCTYPE_COLS = ["code", "name", "description", "synonyms", "keywords", "active",
                "formats", "max_mb", "max_count", "feeds_sections"]
INDUSTRY_COLS = ["sector_code", "sector_name", "industry_code", "industry_name"]
PROMPT_COLS = ["section_code", "section_name", "scope", "prompt_text",
               "source_doc_types", "uses_industry_kpis", "uses_external_context",
               "rendering_hints", "model", "temperature", "max_tokens"]
KPI_COLS = ["industry_code", "kpi_code", "kpi_name", "definition", "unit",
            "polarity", "benchmark", "sections"]
TEMPLATE_COLS = ["key", "name", "segment", "relationship", "template_instructions",
                 "required_doc_types"]
TEMPLATE_SECTION_COLS = ["template_key", "order", "section_code", "mandatory",
                         "include_if_doctype", "length_guidance", "fixed_format"]

EXAMPLES: dict[str, list[list[Any]]] = {
    "doctypes": [["example_audited_financials", "Audited financials",
                  "Annual audited accounts", "annual report|audited accounts",
                  "balance sheet|profit and loss|cash flow", True,
                  "pdf|docx", 25, 10, ""]],
    "industries": [["mfg", "Manufacturing", "example_steel", "Steel"]],
    "prompts": [["example_exec_summary", "Executive Summary", "section",
                 "Draft the executive summary for {{borrower_name}} using "
                 "{{doc:audited_financials}}.",
                 "audited_financials", False, False, "", "", "", ""]],
    "kpi_sets": [["example_steel", "ebitda_per_tonne", "EBITDA per tonne",
                  "Operating profit over tonnes sold", "INR/t", "higher_better",
                  "4500", "industry_analysis|financial_analysis"]],
    "templates": [["example_corp_template", "Corporate CAM - Example", "corporate",
                   "etb", "House style: UK English, amounts in INR crore.",
                   "audited_financials|bank_statements"]],
    "template_sections": [["example_corp_template", 1, "exec_summary", True, "",
                           "250 words", True]],
}

_SHEET_COLS = {
    "doctypes": DOCTYPE_COLS, "industries": INDUSTRY_COLS, "prompts": PROMPT_COLS,
    "kpi_sets": KPI_COLS, "templates": TEMPLATE_COLS,
    "template_sections": TEMPLATE_SECTION_COLS,
}

README_LINES = [
    "CAM master configuration — bulk upload template",
    "",
    "Fill in one row per master entry, then upload this workbook on the "
    "Masters > Bulk import screen (or POST it to /api/masters/bulk-upload).",
    "",
    "Every imported entry lands as a DRAFT — it must still be submitted and "
    "approved by a different admin (maker-checker) before it takes effect.",
    "",
    "Conventions:",
    "  - List cells (synonyms, keywords, source_doc_types, sections, ...) use "
    "'|' to separate values, e.g.  balance sheet|profit and loss",
    "  - Boolean cells accept TRUE/FALSE.",
    "  - A row is IGNORED if its first (key) cell is blank, starts with '#', "
    "or starts with 'example' — so the example rows below are safe to keep.",
    "",
    "Dependency order is handled for you on import: document types and "
    "industries first, then prompts, KPI sets, and templates last.",
    "",
    "Sheets:",
    "  doctypes           — one row per document type (key column: code)",
    "  industries         — one row per industry (key column: industry_code)",
    "  prompts            — one row per section/agent prompt (key: section_code)",
    "  kpi_sets           — one row per KPI; rows sharing an industry_code form "
    "one KPI set",
    "  templates          — one row per CAM template (key column: key)",
    "  template_sections  — one row per template section, linked by template_key",
]


# ------------------------------------------------------------------ build
def build_template_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "README"
    for line in README_LINES:
        ws.append([line])
    ws.column_dimensions["A"].width = 100

    for sheet, cols in _SHEET_COLS.items():
        w = wb.create_sheet(sheet)
        w.append(cols)
        for c in range(1, len(cols) + 1):
            w.cell(row=1, column=c).font = Font(bold=True)
            w.column_dimensions[w.cell(row=1, column=c).column_letter].width = 22
        for row in EXAMPLES.get(sheet, []):
            w.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------ parse helpers
def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _list(v: Any) -> list[str]:
    text = _s(v).replace("\n", "|").replace(",", "|")
    return [x.strip() for x in text.split("|") if x.strip()]


def _bool(v: Any, default: bool = False) -> bool:
    if v is None or _s(v) == "":
        return default
    if isinstance(v, bool):
        return v
    return _s(v).lower() in ("true", "1", "yes", "y")


def _skip_key(v: Any) -> bool:
    k = _s(v).lower()
    return (not k) or k.startswith("#") or k.startswith("example")


def _num(v: Any, kind: str):
    """int/float coercion; raises ValueError with a friendly message."""
    if v is None or _s(v) == "":
        return None
    if isinstance(v, bool):
        raise ValueError(f"expected a {kind}, got a boolean")
    try:
        return int(float(v)) if kind == "int" else float(v)
    except (TypeError, ValueError, OverflowError):
        # OverflowError: int(float("1e400")) -> int(inf); re-raise as a per-row
        # ValueError the row handlers already convert to a reported error.
        raise ValueError(f"'{_s(v)}' is not a valid {kind}")


def _sheet_rows(wb, sheet: str, cols: list[str], key_col: str | None = None,
                errors: list | None = None):
    """Yield (excel_row_no, {col: value}) for a sheet, mapping by header name.
    Missing sheet -> nothing. If key_col is given, the sheet has populated data
    rows, but that column is absent from the header, append a sheet-level error
    (so a recased/renamed/deleted key header is not silent data loss) and yield
    nothing."""
    if sheet not in wb.sheetnames:
        return
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    header = [_s(h) for h in rows[0]]
    data = [(n, raw) for n, raw in enumerate(rows[1:], start=2)
            if not (raw is None or all(c is None or _s(c) == "" for c in raw))]
    if key_col and errors is not None and data and key_col not in header:
        errors.append({"sheet": sheet, "row": 1,
                       "message": f"key column '{key_col}' not found in the header row; "
                                  "column headers must match the template exactly"})
        return
    idx = {name: header.index(name) for name in cols if name in header}
    for excel_row, raw in data:
        rowd = {name: (raw[i] if i < len(raw) else None) for name, i in idx.items()}
        yield excel_row, rowd


# ------------------------------------------------------------------ parse
def parse_workbook(raw: bytes) -> tuple[list[dict], list[dict]]:
    """Return (entries, errors).

    entries: [{"mtype","key","payload"}]  (unsorted; the importer orders them)
    errors:  [{"sheet","row","message"}]  (rows that could not be shaped)
    """
    errors: list[dict] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        return [], [{"sheet": "-", "row": 0,
                     "message": "file is not a readable .xlsx workbook"}]

    entries: list[dict] = []
    try:
        entries += _parse_doctypes(wb, errors)
        entries += _parse_industries(wb, errors)
        entries += _parse_prompts(wb, errors)
        entries += _parse_kpi_sets(wb, errors)
        entries += _parse_templates(wb, errors)
    finally:
        wb.close()
    return entries, errors


def _parse_doctypes(wb, errors) -> list[dict]:
    out = []
    for rn, r in _sheet_rows(wb, "doctypes", DOCTYPE_COLS, "code", errors):
        if _skip_key(r.get("code")):
            continue
        try:
            fc = {}
            formats = _list(r.get("formats"))
            if formats:
                fc["formats"] = formats
            mb, count = _num(r.get("max_mb"), "int"), _num(r.get("max_count"), "int")
            if mb is not None:
                fc["max_mb"] = mb
            if count is not None:
                fc["max_count"] = count
        except ValueError as exc:
            errors.append({"sheet": "doctypes", "row": rn, "message": str(exc)})
            continue
        payload = {"code": _s(r.get("code")), "name": _s(r.get("name")),
                   "description": _s(r.get("description")),
                   "synonyms": _list(r.get("synonyms")),
                   "keywords": _list(r.get("keywords")),
                   "active": _bool(r.get("active"), True),
                   "feeds_sections": _list(r.get("feeds_sections"))}
        if fc:
            payload["file_constraints"] = fc
        out.append({"mtype": "doctype", "key": payload["code"], "payload": payload})
    return out


def _parse_industries(wb, errors) -> list[dict]:
    out = []
    for _rn, r in _sheet_rows(wb, "industries", INDUSTRY_COLS, "industry_code", errors):
        if _skip_key(r.get("industry_code")):
            continue
        payload = {"sector_code": _s(r.get("sector_code")),
                   "sector_name": _s(r.get("sector_name")),
                   "industry_code": _s(r.get("industry_code")),
                   "industry_name": _s(r.get("industry_name"))}
        out.append({"mtype": "industry", "key": payload["industry_code"], "payload": payload})
    return out


def _parse_prompts(wb, errors) -> list[dict]:
    out = []
    for rn, r in _sheet_rows(wb, "prompts", PROMPT_COLS, "section_code", errors):
        if _skip_key(r.get("section_code")):
            continue
        try:
            overrides = {}
            if _s(r.get("model")):
                overrides["model"] = _s(r.get("model"))
            temp = _num(r.get("temperature"), "float")
            if temp is not None:
                overrides["temperature"] = temp
            mt = _num(r.get("max_tokens"), "int")
            if mt is not None:
                overrides["max_tokens"] = mt
        except ValueError as exc:
            errors.append({"sheet": "prompts", "row": rn, "message": str(exc)})
            continue
        payload = {"section_code": _s(r.get("section_code")),
                   "section_name": _s(r.get("section_name")),
                   "scope": _s(r.get("scope")) or "section",
                   "prompt_text": _s(r.get("prompt_text")),
                   "source_doc_types": _list(r.get("source_doc_types")),
                   "uses_industry_kpis": _bool(r.get("uses_industry_kpis")),
                   "uses_external_context": _bool(r.get("uses_external_context")),
                   "rendering_hints": _s(r.get("rendering_hints"))}
        if overrides:
            payload["model_overrides"] = overrides
        out.append({"mtype": "prompt", "key": payload["section_code"], "payload": payload})
    return out


def _parse_kpi_sets(wb, errors) -> list[dict]:
    grouped: dict[str, list[dict]] = {}
    for rn, r in _sheet_rows(wb, "kpi_sets", KPI_COLS, "industry_code", errors):
        if _skip_key(r.get("industry_code")):
            continue
        code = _s(r.get("kpi_code"))
        if not code:
            errors.append({"sheet": "kpi_sets", "row": rn, "message": "kpi_code is required"})
            continue
        grouped.setdefault(_s(r.get("industry_code")), []).append({
            "code": code, "name": _s(r.get("kpi_name")),
            "definition": _s(r.get("definition")), "unit": _s(r.get("unit")),
            "polarity": _s(r.get("polarity")),
            "benchmark": _s(r.get("benchmark")) or None,
            "sections": _list(r.get("sections"))})
    return [{"mtype": "kpi_set", "key": ind, "payload": {"industry_code": ind, "kpis": kpis}}
            for ind, kpis in grouped.items()]


def _parse_templates(wb, errors) -> list[dict]:
    sections: dict[str, list[dict]] = {}
    for rn, r in _sheet_rows(wb, "template_sections", TEMPLATE_SECTION_COLS,
                             "template_key", errors):
        if _skip_key(r.get("template_key")):
            continue
        try:
            order = _num(r.get("order"), "int")
        except ValueError as exc:
            errors.append({"sheet": "template_sections", "row": rn, "message": str(exc)})
            continue
        sections.setdefault(_s(r.get("template_key")), []).append({
            "order": order if order is not None else 1,
            "section_code": _s(r.get("section_code")),
            "mandatory": _bool(r.get("mandatory"), True),
            "include_if_doctype": _s(r.get("include_if_doctype")) or None,
            "length_guidance": _s(r.get("length_guidance")),
            "fixed_format": _bool(r.get("fixed_format"))})

    out, consumed = [], set()
    for _rn, r in _sheet_rows(wb, "templates", TEMPLATE_COLS, "key", errors):
        if _skip_key(r.get("key")):
            continue
        key = _s(r.get("key"))
        consumed.add(key)
        payload = {"name": _s(r.get("name")), "segment": _s(r.get("segment")),
                   "relationship": _s(r.get("relationship")),
                   "template_instructions": _s(r.get("template_instructions")),
                   "sections": sorted(sections.get(key, []), key=lambda s: s["order"]),
                   "required_doc_types": _list(r.get("required_doc_types"))}
        out.append({"mtype": "template", "key": key, "payload": payload})

    # section rows keyed to a template that never appears -> flag, don't drop silently
    for orphan in sorted(set(sections) - consumed):
        errors.append({"sheet": "template_sections", "row": 0,
                       "message": f"template_key '{orphan}' has section rows but no matching "
                                  "row in the templates sheet"})
    return out
